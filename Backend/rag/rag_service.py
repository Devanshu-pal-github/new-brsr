
# Table Extraction Logic for RAG
import os

# Directory to store all FAISS index folders
FAISS_INDEX_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'faiss_indexes')
os.makedirs(FAISS_INDEX_DIR, exist_ok=True)
import re
import os

# Suppress TensorFlow warnings for cleaner startup
os.environ.setdefault('TF_CPP_MIN_LOG_LEVEL', '2')  # Suppress TensorFlow info/warning logs
os.environ.setdefault('TF_ENABLE_ONEDNN_OPTS', '0')  # Disable oneDNN for consistent results

try:
    from pint import UnitRegistry
    ureg = UnitRegistry()
except ImportError:
    ureg = None  # If pint is not installed, skip unit conversion

# Standard library imports
import os
import uuid
import datetime
import shutil
import re
import time
import hashlib
from typing import List
from io import BytesIO

# Third-party imports
from dotenv import load_dotenv
from pymongo import MongoClient
from langchain_community.vectorstores import FAISS
from langchain.text_splitter import RecursiveCharacterTextSplitter
from langchain_huggingface import HuggingFaceEmbeddings
from pypdf import PdfReader
from google import genai
from google.genai import types

# For Word and Excel support
try:
    import docx
    from openpyxl import load_workbook
    DOCX_AVAILABLE = True
    EXCEL_AVAILABLE = True
except ImportError:
    DOCX_AVAILABLE = False
    EXCEL_AVAILABLE = False
    print("⚠️ [RAG] docx or openpyxl not available. Word/Excel support disabled.")

# Load environment variables
load_dotenv()
GEMINI_API_KEY_RAG = os.getenv("GEMINI_API_KEY_RAG")
GEMINI_MODEL = "gemini-2.0-flash"

# Global variables for caching and rate limiting
API_CACHE = {}
LAST_API_CALL = {}
MIN_API_INTERVAL = 2.0  # Minimum 2 seconds between API calls per endpoint
CACHE_DURATION = 3600  # Cache for 1 hour

def get_cache_key(content: str, question: str) -> str:
    """Generate a cache key for the request"""
    combined = f"{content}|{question}"
    return hashlib.md5(combined.encode()).hexdigest()

def is_api_rate_limited(endpoint: str) -> bool:
    """Check if we need to wait before making an API call"""
    if endpoint not in LAST_API_CALL:
        return False
    
    time_since_last = time.time() - LAST_API_CALL[endpoint]
    return time_since_last < MIN_API_INTERVAL

def wait_for_rate_limit(endpoint: str):
    """Wait if necessary to respect rate limits"""
    if endpoint in LAST_API_CALL:
        time_since_last = time.time() - LAST_API_CALL[endpoint]
        if time_since_last < MIN_API_INTERVAL:
            sleep_time = MIN_API_INTERVAL - time_since_last
            print(f"Rate limiting: waiting {sleep_time:.2f} seconds before API call")
            time.sleep(sleep_time)

def update_last_api_call(endpoint: str):
    """Update the timestamp of the last API call"""
    LAST_API_CALL[endpoint] = time.time()

def get_cached_response(cache_key: str):
    """Get cached response if available and not expired"""
    if cache_key in API_CACHE:
        cached_data = API_CACHE[cache_key]
        if time.time() - cached_data['timestamp'] < CACHE_DURATION:
            print(f"Using cached response for key: {cache_key[:8]}...")
            return cached_data['response']
        else:
            # Remove expired cache
            del API_CACHE[cache_key]
    return None

def cache_response(cache_key: str, response):
    """Cache the API response"""
    API_CACHE[cache_key] = {
        'response': response,
        'timestamp': time.time()
    }
    print(f"Cached response for key: {cache_key[:8]}...")
MONGODB_URL = os.getenv("MONGODB_URL")
DB_NAME = os.getenv("DB_NAME")

# Global variables for lazy loading
_embeddings = None
_text_splitter = None
_genai_client = None

# Cache for API responses to avoid duplicate calls
_api_response_cache = {}
_last_api_call_time = 0
_min_api_call_interval = 1.0  # Minimum 1 second between API calls

def get_embeddings():
    """Lazy load embeddings to avoid slow startup times."""
    global _embeddings
    if _embeddings is None:
        print("🔍 [RAG] Loading embeddings model (first time only)...")
        _embeddings = HuggingFaceEmbeddings(model_name="sentence-transformers/all-MiniLM-L6-v2")
        print("🔍 [RAG] Embeddings model loaded successfully")
    return _embeddings

def get_text_splitter():
    """Lazy load text splitter."""
    global _text_splitter
    if _text_splitter is None:
        _text_splitter = RecursiveCharacterTextSplitter(chunk_size=1000, chunk_overlap=200)
    return _text_splitter

def get_genai_client():
    """Lazy load Gemini client."""
    global _genai_client
    if _genai_client is None:
        _genai_client = genai.Client(api_key=GEMINI_API_KEY_RAG)
    return _genai_client

async def get_file_metadata(file_id: str, db):
    return await db.rag_files.find_one({"_id": file_id})

def get_vector_store(file_id: str):
    faiss_path = os.path.join(FAISS_INDEX_DIR, f"faiss_index_{file_id}")
    embeddings = get_embeddings()  # Use lazy-loaded embeddings
    return FAISS.load_local(faiss_path, embeddings, allow_dangerous_deserialization=True)

def retrieve_relevant_chunks(file_id: str, question: str, k: int = 3):
    print(f"🔍 [RAG] Retrieving chunks for file_id: {file_id}, question: {question[:100]}...")
    try:
        vector_store = get_vector_store(file_id)
        docs = vector_store.similarity_search(question, k=k)
        print(f"🔍 [RAG] Found {len(docs)} relevant chunks")
        for i, doc in enumerate(docs):
            print(f"🔍 [RAG] Chunk {i+1}: {doc.page_content[:100]}...")
        return docs
    except Exception as e:
        print(f"❌ [RAG] Error retrieving chunks: {e}")
        return []

def ask_gemini_with_context(context: str, question: str) -> str:
    print(f"🔍 [RAG] Asking Gemini with context length: {len(context)}")
    
    # Generate cache key
    cache_key = get_cache_key(context, question)
    
    # Check cache first
    cached_response = get_cached_response(cache_key)
    if cached_response:
        return cached_response
    
    prompt = (
        "You are a chatbot that answers questions strictly based on the provided document. "
        "Do not use any external knowledge or assumptions.\n\n"
        f"Document content:\n{context}\n\n"
        f"Question: {question}\n\n"
        "Answer:"
    )
    print(f"🔍 [RAG] Full prompt: {prompt[:500]}...")
    
    try:
        # Apply rate limiting
        wait_for_rate_limit("gemini_qa")
        
        generate_content_config = types.GenerateContentConfig(response_mime_type="text/plain")
        genai_client = get_genai_client()  # Use lazy-loaded client
        
        # Update rate limit tracker
        update_last_api_call("gemini_qa")
        
        response = genai_client.models.generate_content(
            model=GEMINI_MODEL,
            contents=prompt,
            config=generate_content_config,
        )
        if not response or not hasattr(response, 'text'):
            print("❌ [RAG] No response from Gemini API")
            return "No response from Gemini API."
        
        result = str(response.text)
        print(f"🔍 [RAG] Gemini response: {result}")
        
        # Cache the successful response
        cache_response(cache_key, result)
        
        return result
    except Exception as e:
        error_msg = str(e)
        print(f"❌ [RAG] Error calling Gemini: {error_msg}")
        
        # Check for specific API quota errors
        if "429" in error_msg or "RESOURCE_EXHAUSTED" in error_msg or "quota" in error_msg.lower():
            return "API_QUOTA_EXCEEDED: Please check your Gemini API quota and billing details."
        elif "403" in error_msg or "permission" in error_msg.lower():
            return "API_PERMISSION_ERROR: Please check your Gemini API key permissions."
        else:
            return f"API_ERROR: {error_msg}"

def extract_text_from_file(file_bytes: bytes, filename: str) -> str:
    """Extract text from PDF, Word, or Excel files."""
    print(f"🔍 [RAG] Extracting text from file: {filename}")
    
    file_ext = filename.lower().split('.')[-1]
    text = ""
    
    try:
        if file_ext == 'pdf':
            pdf_reader = PdfReader(BytesIO(file_bytes))
            for page_num, page in enumerate(pdf_reader.pages):
                page_text = page.extract_text() or ""
                text += page_text
                print(f"🔍 [RAG] Extracted {len(page_text)} chars from PDF page {page_num + 1}")
        
        elif file_ext in ['doc', 'docx'] and DOCX_AVAILABLE:
            doc = docx.Document(BytesIO(file_bytes))
            for paragraph in doc.paragraphs:
                text += paragraph.text + "\n"
            print(f"🔍 [RAG] Extracted {len(text)} chars from Word document")
        
        elif file_ext in ['xls', 'xlsx'] and EXCEL_AVAILABLE:
            workbook = load_workbook(BytesIO(file_bytes), data_only=True)
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text += f"Sheet: {sheet_name}\n"
                for row in sheet.iter_rows(values_only=True):
                    row_text = " | ".join([str(cell) if cell is not None else "" for cell in row])
                    text += row_text + "\n"
            print(f"🔍 [RAG] Extracted {len(text)} chars from Excel document")
        
        else:
            raise ValueError(f"Unsupported file type: {file_ext}")
        
        print(f"🔍 [RAG] Total extracted text length: {len(text)}")
        print(f"🔍 [RAG] Text preview: {text[:300]}...")
        return text
        
    except Exception as e:
        print(f"❌ [RAG] Error extracting text from {filename}: {e}")
        raise e

async def process_file_and_store(file_bytes: bytes, filename: str, file_size: int, db) -> str:
    """Extract text from any supported file type, split, embed, store in FAISS and MongoDB. Returns file_id."""
    print(f"🔍 [RAG] Processing file: {filename} ({file_size} bytes)")
    
    # Extract text using the universal extractor
    text = extract_text_from_file(file_bytes, filename)
    
    if not text.strip():
        raise ValueError("No text could be extracted from the file")
    
    print(f"🔍 [RAG] Creating documents from text ({len(text)} characters)")
    text_splitter = get_text_splitter()  # Use lazy-loaded text splitter
    documents = text_splitter.create_documents([text])
    print(f"🔍 [RAG] Created {len(documents)} document chunks")
    
    # Create vector store
    embeddings = get_embeddings()  # Use lazy-loaded embeddings
    vector_store = FAISS.from_documents(documents, embeddings)
    file_id = str(uuid.uuid4())
    faiss_path = os.path.join(FAISS_INDEX_DIR, f"faiss_index_{file_id}")
    vector_store.save_local(faiss_path)
    print(f"🔍 [RAG] Saved vector store to: {faiss_path}")

    # Store metadata in MongoDB
    await db.rag_files.insert_one({
        "_id": file_id,
        "filename": filename,
        "content_length": file_size,
        "text_length": len(text),
        "chunk_count": len(documents),
        "faiss_path": faiss_path,
        "upload_date": datetime.datetime.now().timestamp()
    })
    print(f"🔍 [RAG] Stored metadata for file_id: {file_id}")

    return file_id


async def delete_file_and_index(file_id: str, db) -> bool:
    file_metadata = await db.rag_files.find_one({"_id": file_id})
    if not file_metadata:
        return False
    faiss_path = os.path.join(FAISS_INDEX_DIR, f"faiss_index_{file_id}")
    try:
        shutil.rmtree(faiss_path)
    except Exception:
        pass  # Ignore errors if directory does not exist
    await db.rag_files.delete_one({"_id": file_id})
    return True


async def extract_table_values(file_id: str, table_metadata: dict, question: str):
    """
    Extracts values for each editable cell in the table using RAG and Gemini.
    Returns: { 'suggested_values': {rowIdx: {colKey: value, ...}, ...}, 'unit_warnings': [ ... ] }
    """
    print(f"🔍 [RAG] Starting table extraction for file_id: {file_id}")
    print(f"🔍 [RAG] Question: {question}")
    print(f"🔍 [RAG] Table metadata keys: {list(table_metadata.keys())}")
    
    # Only process if rows and columns exist
    rows = table_metadata.get('rows', [])
    columns = table_metadata.get('columns', [])
    print(f"🔍 [RAG] Found {len(rows)} rows and {len(columns)} columns")
    print(f"🔍 [RAG] All rows:")
    for idx, row in enumerate(rows):
        print(f"  Row {idx}: {row}")
    print(f"🔍 [RAG] All columns:")
    for idx, col in enumerate(columns):
        print(f"  Col {idx}: {col}")
    
    # Identify editable rows (not auto-calculated)
    # Remove HTML tags from parameter text for keyword checking
    import re
    def clean_html(text):
        return re.sub(r'<[^>]+>', '', text)
    
    editable_row_indices = []
    for idx, row in enumerate(rows):
        if row.get('isHeader') or row.get('isSectionHeader'):
            continue
        
        # For dynamic modules, check if row has a 'label' field (row parameter)
        # For environment modules, check 'parameter' field
        param = row.get('label') or row.get('parameter', '')
        param_clean = clean_html(param).lower()
        print(f"🔍 [RAG] Row {idx}: '{param}' -> cleaned: '{param_clean}'")
        
        # Check if it's auto-calculated (very specific checks)
        is_auto_calc = (
            # Intensity calculations (per unit measures)
            ('intensity' in param_clean and ('per' in param_clean or '/' in param_clean)) or
            ('per rupee' in param_clean) or
            ('per ton' in param_clean) or
            ('/₹' in param_clean) or
            ('/ton' in param_clean) or
            # Sum/Total rows that explicitly show addition (A+B+C pattern)
            (('total' in param_clean or 'sum' in param_clean) and 
             ('+' in param_clean or 'a+b+c' in param_clean))
        )
        
        if not is_auto_calc:
            editable_row_indices.append(idx)
            print(f"🔍 [RAG] Row {idx} marked as editable")
        else:
            print(f"🔍 [RAG] Row {idx} marked as auto-calculated")
    
    print(f"🔍 [RAG] Editable row indices: {editable_row_indices}")
    
    # Identify year columns (e.g., 'current_year', 'previous_year')
    year_col_keys = [col.get('key') for col in columns if col.get('key') not in ('parameter', 'unit')]
    # If dynamicYear, fallback to all except parameter/unit
    if not year_col_keys:
        year_col_keys = [col.get('key') for col in columns if col.get('key') not in ('parameter', 'unit')]
    print(f"🔍 [RAG] Year column keys: {year_col_keys}")
    
    suggested_values = {}
    unit_warnings = []
    
    # A.1 Initialize context cache to avoid redundant chunk retrievals
    context_cache = {}
    
    # A.2 Get document context once for the entire table extraction
    print(f"🔍 [RAG] Retrieving document context for table extraction...")
    main_context_queries = [
        question,  # Use the user's question
        "emissions air pollutants",  # Generic search for emissions data
        "NOx SOx PM VOC POP HAP",  # Search for specific pollutants
        "kg/year g/year",  # Search by units
    ]
    
    # Build comprehensive context from all queries
    all_docs = []
    for search_query in main_context_queries:
        cache_key = f"{file_id}_{search_query}"
        if cache_key not in context_cache:
            docs = retrieve_relevant_chunks(file_id, search_query, k=3)
            context_cache[cache_key] = docs
        else:
            docs = context_cache[cache_key]
        all_docs.extend(docs)
    
    # Remove duplicates and get unique content
    unique_content = []
    seen_content = set()
    for doc in all_docs:
        if doc.page_content not in seen_content:
            unique_content.append(doc.page_content)
            seen_content.add(doc.page_content)
    
    # Use the same context for all cells to avoid redundant API calls
    shared_context = "\n".join(unique_content)
    print(f"🔍 [RAG] Built shared context: {len(shared_context)} chars from {len(unique_content)} unique chunks")
    print(f"🔍 [RAG] Context preview: {shared_context[:300]}...")
    
    # Initialize suggested_values structure
    for row_idx in editable_row_indices:
        suggested_values[str(row_idx)] = {}
        for col_key in year_col_keys:
            suggested_values[str(row_idx)][col_key] = ""
    
    # C. SINGLE API CALL APPROACH - Extract all cell values in one go
    print(f"🔍 [RAG] Making single API call to extract all table cell values...")
    print(f"🔍 [RAG] This replaces {len(editable_row_indices)} * {len(year_col_keys)} = {len(editable_row_indices) * len(year_col_keys)} individual API calls with just 1 call!")
    
    # Build comprehensive question for all parameter-column combinations
    all_params = []
    all_columns = []
    cell_combinations = []
    
    for row_idx in editable_row_indices:
        row = rows[row_idx]
        param = row.get('label') or row.get('parameter', '')
        param_clean = clean_html(param)
        all_params.append(param_clean)
        print(f"🔍 [RAG] Parameter {len(all_params)}: {param_clean}")
    
    for col in columns:
        if col.get('key') in year_col_keys:
            all_columns.append({
                'key': col.get('key'),
                'label': col.get('label', ''),
            })
            print(f"🔍 [RAG] Column: {col.get('key')} - {col.get('label', '')}")
    
    # Create all parameter-column combinations
    for param_idx, param in enumerate(all_params):
        for col in all_columns:
            cell_combinations.append({
                'param_idx': param_idx,
                'param': param,
                'col_key': col['key'],
                'col_label': col['label']
            })
    
    print(f"🔍 [RAG] Total cell combinations to extract: {len(cell_combinations)}")
    
    # Create a single comprehensive question that asks for all cell values
    comprehensive_question = f"""From the document, extract the specific values for each parameter-column combination in this retirement benefits table.

The table has the following structure:
Parameters (rows): {', '.join(all_params)}
Columns: {', '.join([f"{col['key']} ({col['label']})" for col in all_columns])}

Please extract the value for each cell by finding the intersection of parameter and column. Return the values in this exact format:

"""
    
    # Add each cell combination to the question
    for i, combo in enumerate(cell_combinations):
        comprehensive_question += f"{i+1}. {combo['param']} x {combo['col_label']}: [value]\n"
    
    comprehensive_question += f"""
Instructions:
- Look for the table structure in the document showing retirement benefits data
- Find the intersection value for each parameter (row) and column combination
- If a cell shows "Y", return "Y"
- If a cell shows "N" or "N.A.", return "N"
- If a cell shows a percentage like "95%", return just the number "95"
- If a cell is empty or not found, return an empty line
- Return exactly {len(cell_combinations)} values, one per line, in the order listed above

Example format:
95
90
Y
95
90
Y
100
95
Y
100
95
Y
..."""
    
    print(f"🔍 [RAG] Making single comprehensive API call...")
    print(f"🔍 [RAG] Question preview: {comprehensive_question[:800]}...")
    
    # SINGLE API CALL for all cell combinations
    answer = ask_gemini_with_context(shared_context, comprehensive_question)
    print(f"🔍 [RAG] ✅ Single API call completed!")
    print(f"🔍 [RAG] Raw answer: {answer}")
    
    # Check if the answer contains an error
    if any(error_indicator in answer for error_indicator in [
        "API_QUOTA_EXCEEDED", "API_PERMISSION_ERROR", "API_ERROR", 
        "Error calling Gemini:", "429 RESOURCE_EXHAUSTED", "quota exceeded"
    ]):
        print(f"❌ [RAG] Gemini API error detected in comprehensive call")
        print(f"❌ [RAG] Error details: {answer}")
        print(f"❌ [RAG] All table values will remain empty due to API error")
        # Keep all values empty if API fails (already initialized above)
    else:
        # Parse the comprehensive answer
        print(f"🔍 [RAG] Parsing comprehensive answer...")
        answer_lines = answer.strip().split('\n')
        
        # Extract all values from the answer
        extracted_values = []
        for i, line in enumerate(answer_lines):
            line = line.strip()
            print(f"🔍 [RAG] Parsing line {i+1}: '{line}'")
            
            if not line:
                extracted_values.append("")
                print(f"🔍 [RAG] Empty line -> empty value")
                continue
            
            # Look for values in brackets first [95], [Yes], [N.A.], etc.
            bracket_match = re.search(r'\[([^\]]+)\]', line)
            if bracket_match:
                bracket_value = bracket_match.group(1).strip()
                print(f"🔍 [RAG] Found bracketed value: '{bracket_value}'")
                
                # Handle different value types
                if bracket_value.upper() in ['Y', 'YES']:
                    extracted_values.append("Y")
                    print(f"🔍 [RAG] Converted to: Y")
                elif bracket_value.upper() in ['N', 'NO', 'N.A.', 'NA']:
                    extracted_values.append("N")
                    print(f"🔍 [RAG] Converted to: N")
                else:
                    # Try to extract number from bracketed value
                    numbers = re.findall(r'[\d,]+(?:\.[\d]+)?', bracket_value)
                    if numbers:
                        value = numbers[0].replace(',', '')
                        try:
                            float(value)  # Validate it's a number
                            extracted_values.append(value)
                            print(f"🔍 [RAG] Converted to number: {value}")
                        except ValueError:
                            extracted_values.append("")
                            print(f"🔍 [RAG] Invalid number format -> empty value")
                    else:
                        # Check for nil/zero indicators in bracket
                        if any(word in bracket_value.lower() for word in ['nil', 'zero', 'not applicable', 'n/a', 'none']):
                            extracted_values.append("0")
                            print(f"🔍 [RAG] Found nil/zero indicator -> 0")
                        else:
                            extracted_values.append("")
                            print(f"🔍 [RAG] No recognizable value in brackets -> empty value")
                continue
            
            # Fallback: Handle different value types without brackets
            if line.upper() in ['Y', 'YES']:
                extracted_values.append("Y")
                print(f"🔍 [RAG] Found Y/Yes -> Y")
            elif line.upper() in ['N', 'NO', 'N.A.', 'NA']:
                extracted_values.append("N")
                print(f"🔍 [RAG] Found N/No/N.A. -> N")
            else:
                # Find numbers in the line (avoid line numbers at the beginning)
                # Look for numbers that are not at the very start of the line
                numbers = re.findall(r'(?:^|\s)(\d+(?:\.\d+)?)(?:\s|$|%)', line)
                if not numbers:
                    # If no spaced numbers found, try any numbers
                    numbers = re.findall(r'[\d,]+(?:\.[\d]+)?', line)
                
                if numbers:
                    # Take the last number found (more likely to be the value, not line number)
                    value = numbers[-1].replace(',', '')
                    try:
                        float(value)  # Validate it's a number
                        extracted_values.append(value)
                        print(f"🔍 [RAG] Found number: {value}")
                    except ValueError:
                        extracted_values.append("")
                        print(f"🔍 [RAG] Invalid number format -> empty value")
                else:
                    # Check for nil/zero indicators
                    if any(word in line.lower() for word in ['nil', 'zero', 'not applicable', 'n/a', 'none']):
                        extracted_values.append("0")
                        print(f"🔍 [RAG] Found nil/zero indicator -> 0")
                    else:
                        extracted_values.append("")
                        print(f"🔍 [RAG] No recognizable value -> empty value")
        
        print(f"🔍 [RAG] Final extracted values: {extracted_values}")
        
        # Assign extracted values to the specific cells
        for i, combo in enumerate(cell_combinations):
            if i < len(extracted_values):
                value = extracted_values[i]
                param_idx = combo['param_idx']
                row_idx = editable_row_indices[param_idx]
                col_key = combo['col_key']
                
                print(f"🔍 [RAG] Assigning value '{value}' to [{row_idx}][{col_key}] (Parameter: {combo['param']}, Column: {combo['col_label']})")
                suggested_values[str(row_idx)][col_key] = value
            else:
                print(f"🔍 [RAG] No value available for combination {i+1}")
                # Values already initialized as empty above
    
    # Add parameter names to the result for better display
    parameters_info = {}
    for row_idx in editable_row_indices:
        row = rows[row_idx]
        # Handle both dynamic module format ('label') and environment module format ('parameter')
        param = row.get('label') or row.get('parameter', '')
        unit = row.get('unit', '').strip()
        # Clean HTML tags for display
        param_clean = clean_html(param)
        parameters_info[str(row_idx)] = {
            'parameter': param_clean,
            'unit': unit
        }
    
    result = {
        "suggested_values": suggested_values, 
        "unit_warnings": unit_warnings,
        "parameters_info": parameters_info
    }
    print(f"🔍 [RAG] Final result: {result}")
    return result
